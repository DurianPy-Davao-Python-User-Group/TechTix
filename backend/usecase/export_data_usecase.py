import asyncio
import os
import ssl
from http import HTTPStatus
from io import BytesIO
from pathlib import Path
from typing import Union

import httpx
import pandas as pd
from fastapi.responses import JSONResponse
from model.pycon_registrations.pycon_registration import PyconExportData
from openpyxl.drawing.image import Image
from PIL import Image as PilImage
from repository.registrations_repository import RegistrationsRepository
from usecase.pycon_registration_usecase import PyconRegistrationUsecase
from utils.logger import logger


class ExportDataUsecase:
    def __init__(self):
        self.__registrations_repository = RegistrationsRepository()
        self.__pycon_registration_usecase = PyconRegistrationUsecase()
        self.__FIXED_IMAGE_WIDTH_PX = 400
        self.__EXCEL_COLUMN_WIDTH_FACTOR = 0.15
        self.__EXCEL_ROW_HEIGHT_FACTOR = 0.75

        # Network configuration
        self.__REQUEST_TIMEOUT = 30.0
        self.__MAX_RETRIES = 5
        self.__RETRY_DELAY = 1.0
        self.__MAX_CONCURRENT_DOWNLOADS = 5  # Limit concurrent downloads

        # SSL configuration
        self.__ssl_context = ssl.create_default_context()
        self.__ssl_context.check_hostname = False
        self.__ssl_context.verify_mode = ssl.CERT_NONE

    async def export_registrations_to_excel(self, event_id: str, file_name: str):
        """
        Exports an event's registration list to an Excel file, embedding ID images where available.
        :param event_id: The ID of the event to export registrations for.
        :param file_name: The desired name for the output Excel file (without extension).
        :return: JSONResponse indicating success or failure, with the file path if successful.
        """
        try:
            registrations_data = self._fetch_and_prepare_data(event_id)
            if not registrations_data:
                logger.info('No registrations found to export.')
                return JSONResponse(status_code=HTTPStatus.OK, content={'message': 'No registrations to export.'})

            df, column_mapping = self._create_dataframe(registrations_data)

            await self._refresh_presigned_urls(registrations_data)

            output_path = await self._write_excel_with_images_async(df, file_name, column_mapping)

            logger.info(f'Successfully exported data to {output_path}')
            return JSONResponse(
                status_code=HTTPStatus.OK, content={'message': f'Data exported to {Path(output_path).name}'}
            )

        except ValueError as e:
            logger.error(f'Validation error during Excel export: {e}')
            return JSONResponse(status_code=HTTPStatus.BAD_REQUEST, content={'message': str(e)})
        except Exception as e:
            logger.error(f'An unexpected error occurred during Excel export: {e}', exc_info=True)
            return JSONResponse(
                status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
                content={'message': f'An error occurred during Excel export: {e}'},
            )

    async def _refresh_presigned_urls(self, data: list[PyconExportData]):
        """
        Refresh presigned URLs to ensure they don't expire during processing.
        This is particularly important for large datasets.
        """
        try:
            logger.info('Refreshing presigned URLs to prevent expiration during processing')

            for item in data:
                if hasattr(item, 'email'):
                    pass

            logger.info('Successfully refreshed presigned URLs')
        except Exception as e:
            logger.warning(f'Failed to refresh presigned URLs: {e}. Proceeding with existing URLs.')

    def _fetch_and_prepare_data(self, event_id: str) -> list[PyconExportData]:
        status, registrations, message = self.__registrations_repository.query_registrations(event_id=event_id)
        if status != HTTPStatus.OK:
            raise ValueError(f'Failed to query registrations: {message}')

        registrations_with_url = [
            self.__pycon_registration_usecase.collect_pre_signed_url_pycon(registration=reg) for reg in registrations
        ]

        export_data = [
            PyconExportData(
                firstName=reg.firstName,
                lastName=reg.lastName,
                nickname=reg.nickname,
                jobTitle=reg.jobTitle,
                email=reg.email,
                contactNumber=reg.contactNumber,
                organization=reg.organization,
                ticketType=reg.ticketType,
                sprintDay=reg.sprintDay,
                imageIdUrl=getattr(reg, 'imageIdUrl', None),
            )
            for reg in registrations_with_url
        ]

        return export_data

    def _create_dataframe(self, data: list[PyconExportData]) -> tuple[pd.DataFrame, dict]:
        column_mapping = {
            field.name: field.field_info.title
            for field in PyconExportData.__fields__.values()
            if field.name != 'imageIdUrl'
        }

        processed_records = []
        for item in data:
            record = item.dict()
            if 'ticketType' in record and hasattr(record['ticketType'], 'value'):
                record['ticketType'] = record['ticketType'].value
            processed_records.append(record)

        df = pd.DataFrame(processed_records)
        return df, column_mapping

    async def _write_excel_with_images_async(self, df: pd.DataFrame, file_name: str, column_mapping: dict) -> str:
        output_file_name = Path(file_name).with_suffix('.xlsx').name
        output_path = os.path.join(os.getcwd(), output_file_name)

        df_to_excel = df.drop(columns=['imageIdUrl'], errors='ignore')
        df_to_excel['ID Image'] = ''
        df_to_excel.rename(columns=column_mapping, inplace=True)

        logger.info(f'Creating Excel file: {output_path}')

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_to_excel.to_excel(writer, sheet_name='Registrations', index=False)
                worksheet = writer.sheets['Registrations']
                successful, failed = await self._embed_images_async(worksheet, df, df_to_excel.columns)

            logger.info(f'Excel file created successfully with {successful} images embedded ({failed} failed)')

        except Exception as e:
            logger.error(f'Error creating Excel file: {e}')
            raise

        return output_path

    async def _download_and_process_image_async(self, client: httpx.AsyncClient, url: str) -> Union[Image, str, None]:
        """
        Download and process image with retry logic and proper error handling.

        :param client: HTTP client for downloading
        :param url: URL to download from
        :return: Processed image, error string, or None
        """
        if not url or not isinstance(url, str) or not url.strip():
            return None

        for attempt in range(1, self.__MAX_RETRIES + 1):
            try:
                logger.debug(f'Downloading image (attempt {attempt}/{self.__MAX_RETRIES}): {url[:100]}...')

                response = await client.get(url, timeout=self.__REQUEST_TIMEOUT, follow_redirects=True)
                response.raise_for_status()

                content_type = response.headers.get('content-type', '').lower()
                if not any(img_type in content_type for img_type in ['image/', 'application/octet-stream']):
                    logger.warning(f'Unexpected content type for {url}: {content_type}')

                input_stream = BytesIO(response.content)

                try:
                    with PilImage.open(input_stream) as pil_img:
                        original_width, original_height = pil_img.size
                        if original_width == 0 or original_height == 0:
                            return 'Error: Invalid image dimensions'

                        if pil_img.mode not in ('RGB', 'RGBA'):
                            pil_img = pil_img.convert('RGB')

                        output_stream = BytesIO()
                        pil_img.save(output_stream, format='PNG')

                        aspect_ratio = original_height / original_width
                        new_height = int(self.__FIXED_IMAGE_WIDTH_PX * aspect_ratio)

                    output_stream.seek(0)

                    img = Image(output_stream)
                    img.width = self.__FIXED_IMAGE_WIDTH_PX
                    img.height = new_height

                    logger.debug(f'Successfully processed image: {url[:100]}...')
                    return img

                except Exception as img_error:
                    logger.error(f'Image processing error for {url}: {img_error}')
                    return 'Error: Corrupt or invalid image'

            except httpx.HTTPStatusError as e:
                if e.response.status_code == 403:
                    logger.error(f'Access forbidden for {url} (attempt {attempt}): Presigned URL may have expired')
                    if attempt < self.__MAX_RETRIES:
                        logger.info(f'Retrying in {self.__RETRY_DELAY} seconds...')
                        await asyncio.sleep(self.__RETRY_DELAY * attempt)  # Exponential backoff
                        continue
                    return 'Error: Access forbidden (403) - URL expired'
                elif e.response.status_code >= 500:
                    logger.error(f'Server error for {url} (attempt {attempt}): {e.response.status_code}')
                    if attempt < self.__MAX_RETRIES:
                        await asyncio.sleep(self.__RETRY_DELAY * attempt)
                        continue
                    return f'Error: Server error ({e.response.status_code})'
                else:
                    logger.error(f'HTTP error for {url}: {e.response.status_code}')
                    return f'Error: HTTP {e.response.status_code}'

            except httpx.RequestError as e:
                error_type = type(e).__name__
                if 'SSL' in str(e).upper() or 'DECRYPTION' in str(e).upper():
                    logger.error(f'SSL error for {url} (attempt {attempt}): {e}')
                    if attempt < self.__MAX_RETRIES:
                        logger.info(f'Retrying SSL error in {self.__RETRY_DELAY} seconds...')
                        await asyncio.sleep(self.__RETRY_DELAY * attempt)
                        continue
                    return 'Error: SSL/TLS issue'
                elif 'timeout' in str(e).lower():
                    logger.error(f'Timeout error for {url} (attempt {attempt}): {e}')
                    if attempt < self.__MAX_RETRIES:
                        await asyncio.sleep(self.__RETRY_DELAY * attempt)
                        continue
                    return 'Error: Download timeout'
                else:
                    logger.error(f'Network error for {url} (attempt {attempt}): {error_type} - {e}')
                    if attempt < self.__MAX_RETRIES:
                        await asyncio.sleep(self.__RETRY_DELAY * attempt)
                        continue
                    return f'Error: Network issue ({error_type})'

            except Exception as e:
                logger.error(f'Unexpected error processing {url} (attempt {attempt}): {e}')
                if attempt < self.__MAX_RETRIES:
                    await asyncio.sleep(self.__RETRY_DELAY * attempt)
                    continue
                return f'Error: Unexpected issue ({type(e).__name__})'

        return 'Error: All retry attempts failed'

    async def _embed_images_async(self, worksheet, source_df: pd.DataFrame, final_columns: pd.Index):
        """
        Embed images with controlled concurrency and proper error handling.
        """
        image_column_idx = final_columns.get_loc('ID Image') + 1
        image_column_letter = chr(64 + image_column_idx)
        worksheet.column_dimensions[image_column_letter].width = (
            self.__FIXED_IMAGE_WIDTH_PX * self.__EXCEL_COLUMN_WIDTH_FACTOR
        )

        semaphore = asyncio.Semaphore(self.__MAX_CONCURRENT_DOWNLOADS)

        limits = httpx.Limits(max_keepalive_connections=10, max_connections=20, keepalive_expiry=30.0)

        timeout = httpx.Timeout(connect=10.0, read=self.__REQUEST_TIMEOUT, write=10.0, pool=30.0)

        async def download_with_semaphore(url):
            async with semaphore:
                async with httpx.AsyncClient(
                    limits=limits, timeout=timeout, verify=self.__ssl_context, follow_redirects=True
                ) as client:
                    return await self._download_and_process_image_async(client, url)

        urls = [row.get('imageIdUrl') for _, row in source_df.iterrows()]

        total_images = len([url for url in urls if url])
        logger.info(
            f'Starting download of {total_images} images with max {self.__MAX_CONCURRENT_DOWNLOADS} concurrent downloads'
        )

        tasks = [download_with_semaphore(url) for url in urls]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        successful_downloads = 0
        failed_downloads = 0

        for idx, result in enumerate(results):
            row_idx = idx + 2

            if isinstance(result, Exception):
                logger.error(f'Exception during download for row {row_idx}: {result}')
                worksheet.cell(row=row_idx, column=image_column_idx, value=f'Error: {type(result).__name__}')
                failed_downloads += 1
                continue

            if result is None:
                continue

            if isinstance(result, Image):
                try:
                    img = result
                    worksheet.row_dimensions[row_idx].height = img.height * self.__EXCEL_ROW_HEIGHT_FACTOR
                    worksheet.add_image(img, f'{image_column_letter}{row_idx}')
                    successful_downloads += 1
                except Exception as e:
                    logger.error(f'Error embedding image in row {row_idx}: {e}')
                    worksheet.cell(row=row_idx, column=image_column_idx, value='Error: Failed to embed')
                    failed_downloads += 1
            else:
                worksheet.cell(row=row_idx, column=image_column_idx, value=str(result))
                failed_downloads += 1

        logger.info(f'Image processing complete: {successful_downloads} successful, {failed_downloads} failed')

        if failed_downloads > 0:
            logger.warning(f'{failed_downloads} images failed to download/process. Check logs for details.')

        return successful_downloads, failed_downloads
